
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import sys
from pathlib import Path

# Add the src directory to the Python path to allow for module imports
sys.path.append(str(Path(__file__).parent.parent / 'src'))

from dicom_parser import get_dicom_files, sort_dicom_files, get_plan_data
import pydicom

# --- Helper Functions ---

def get_correct_dwell_times(rtplan_file):
    """Extracts and calculates the correct dwell times from an RTPLAN file."""
    ds = pydicom.dcmread(rtplan_file)
    dwell_times = []
    if hasattr(ds, 'ApplicationSetupSequence'):
        for app_setup in ds.ApplicationSetupSequence:
            if hasattr(app_setup, 'ChannelSequence'):
                for channel in app_setup.ChannelSequence:
                    channel_total_time = float(channel.ChannelTotalTime)
                    final_cumulative_time_weight = float(channel.FinalCumulativeTimeWeight)
                    if hasattr(channel, 'BrachyControlPointSequence'):
                        last_time_weight = 0.0
                        for cp in channel.BrachyControlPointSequence:
                            if hasattr(cp, 'CumulativeTimeWeight'):
                                current_time_weight = float(cp.CumulativeTimeWeight)
                                dwell_time_weight = current_time_weight - last_time_weight
                                if dwell_time_weight > 0:
                                    dwell_time = dwell_time_weight * channel_total_time / final_cumulative_time_weight
                                    dwell_times.append(dwell_time)
                                last_time_weight = current_time_weight
    return dwell_times

def parse_mosaiq_schedule_for_hdr_tx(file_path):
    """
    Parses a Mosaiq schedule to find dates and times for HDR treatments.
    """
    try:
        df = pd.read_excel(file_path)
        hdr_tx_schedule = df[
            df['Activity'].str.contains('HDR', case=False, na=False) & 
            df['Description'].str.contains('tx', case=False, na=False)
        ].copy()
        hdr_tx_schedule = hdr_tx_schedule[~hdr_tx_schedule['Sts'].str.contains('X', na=False)]
        hdr_tx_schedule['Date'] = pd.to_datetime(hdr_tx_schedule['Date'], errors='coerce')
        hdr_tx_schedule['Time'] = hdr_tx_schedule['Time'].astype(str)
        hdr_tx_schedule.dropna(subset=['Date'], inplace=True)
        hdr_tx_schedule['datetime'] = pd.to_datetime(
            hdr_tx_schedule['Date'].dt.strftime('%Y-%m-%d') + ' ' + hdr_tx_schedule['Time']
        )
        return sorted(hdr_tx_schedule['datetime'].tolist())
    except Exception as e:
        print(f"Error parsing Mosaiq schedule file: {e}")
        return []

# --- Main Script ---

def main():
    # --- Define File Paths ---
    mosaiq_schedule_path = r'C:\Users\echorniak\GIT\BrachyD2ccEval\sample_data\Jones Schedule.xlsx.xls'
    template_excel_path = r'C:\Users\echorniak\GIT\BrachyD2ccEval\src\templates\Dwell time decay Worksheet Cylinder.xlsx'
    output_excel_path = r'C:\Users\echorniak\GIT\BrachyD2ccEval\temp_scripts\populated_dwell_time_sheet.xlsx'
    dicom_dir = r'C:\Users\echorniak\GIT\BrachyD2ccEval\sample_data\Jane Doe'
    # -------------------------

    print("1. Parsing Mosaiq schedule for 'HDR: tx' activities...")
    fraction_datetimes = parse_mosaiq_schedule_for_hdr_tx(mosaiq_schedule_path)
    if not fraction_datetimes:
        print("No 'HDR: tx' activities found in the schedule. Exiting.")
        return
    print(f"Found {len(fraction_datetimes)} HDR treatment fractions.")

    # --- Parse DICOM data ---
    print("\n2. Parsing DICOM data...")
    dicom_files = get_dicom_files(dicom_dir)
    sorted_files = sort_dicom_files(dicom_files)
    rtplan_file = sorted_files.get("RTPLAN")
    patient_name = "N/A"
    patient_mrn = "N/A"
    plan_name = "N/A"
    plan_date_str = "N/A"
    dwell_times = []

    if rtplan_file:
        rtplan_dataset = pydicom.dcmread(rtplan_file)
        patient_name = str(rtplan_dataset.PatientName)
        patient_mrn = str(rtplan_dataset.PatientID)
        plan_data = get_plan_data(rtplan_file)
        plan_name = plan_data.get('plan_name', 'N/A')
        plan_date = plan_data.get('plan_date', 'N/A')
        plan_time = plan_data.get('plan_time', 'N/A')
        if plan_date != 'N/A' and plan_time != 'N/A':
            plan_datetime = datetime.strptime(f"{plan_date}{plan_time.split('.')[0]}", "%Y%m%d%H%M%S")
            plan_date_str = plan_datetime.strftime('%Y-%m-%d %H:%M')
        
        print("  - Extracting dwell times...")
        dwell_times = get_correct_dwell_times(rtplan_file)
        print(f"  - Found {len(dwell_times)} dwell times.")

    print(f"  - Patient Name: {patient_name}")
    print(f"  - Patient MRN: {patient_mrn}")
    print(f"  - Plan Name: {plan_name}")
    print(f"  - Plan Date: {plan_date_str}")

    print("\n3. Populating the Excel template...")
    try:
        wb = load_workbook(template_excel_path)
        ws = wb.active

        # --- Populate Header Info ---
        ws['B5'] = patient_name
        ws['B6'] = patient_mrn
        ws['B7'] = plan_name

        # --- Populate Fraction Dates and Times ---
        ws['B11'] = plan_date_str
        fraction_cells = ['C11', 'D11', 'E11', 'F11', 'G11']
        for i, dt in enumerate(fraction_datetimes):
            if i < len(fraction_cells):
                ws[fraction_cells[i]] = dt.strftime('%Y-%m-%d %H:%M')

        # --- Populate Fraction Headers ---
        ws['B9'] = "Plan"
        header_cells = ['C9', 'D9', 'E9', 'F9', 'G9']
        for i in range(len(fraction_datetimes)):
             if i < len(header_cells):
                ws[header_cells[i]] = i + 1
        
        # --- Populate Dwell Times ---
        if dwell_times:
            dwell_time_start_row = 17
            for i, time in enumerate(dwell_times):
                cell_ref = f'B{dwell_time_start_row + i}'
                ws[cell_ref] = time
            print("  - Successfully populated dwell times.")
        else:
            print("  - No dwell times found to populate.")

        wb.save(output_excel_path)
        print(f"\nSuccessfully created populated Excel sheet at: {output_excel_path}")

    except FileNotFoundError:
        print(f"Error: The template file was not found at '{template_excel_path}'.")
    except Exception as e:
        print(f"An error occurred while populating the Excel template: {e}")

if __name__ == "__main__":
    main()
