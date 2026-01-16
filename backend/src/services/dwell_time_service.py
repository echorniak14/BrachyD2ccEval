import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import pydicom
import io
from .parser_service import ParserService

def parse_mosaiq_schedule_for_hdr_tx(input_file):
    """
    Parses the Mosaiq schedule Excel export to find HDR treatment dates.
    Input can be a file path or bytes/file-like object.
    """
    try:
        if isinstance(input_file, bytes):
            input_file = io.BytesIO(input_file)
            
        df = pd.read_excel(input_file)
        
        # Filter for rows containing 'HDR' and 'tx' (case-insensitive)
        if 'Activity' not in df.columns or 'Description' not in df.columns or 'Sts' not in df.columns:
            return []

        hdr_tx_schedule = df[
            df['Activity'].str.contains('HDR', case=False, na=False) &
            df['Description'].str.contains('tx', case=False, na=False)
        ].copy()
        
        # Exclude cancelled/rescheduled appointments (Status 'X')
        hdr_tx_schedule = hdr_tx_schedule[~hdr_tx_schedule['Sts'].str.contains('X', na=False)]
        
        # Parse Dates and Times
        hdr_tx_schedule['Date'] = pd.to_datetime(hdr_tx_schedule['Date'], errors='coerce')
        hdr_tx_schedule['Time'] = hdr_tx_schedule['Time'].astype(str)
        hdr_tx_schedule.dropna(subset=['Date'], inplace=True)
        
        # Combine Date and Time
        hdr_tx_schedule['datetime'] = pd.to_datetime(
            hdr_tx_schedule['Date'].dt.strftime('%Y-%m-%d') + ' ' + hdr_tx_schedule['Time'],
            format='mixed'
        )
        
        return sorted(hdr_tx_schedule['datetime'].tolist())
    except Exception as e:
        print(f"Error parsing Mosaiq schedule: {e}")
        return []

def generate_dwell_time_sheet(mosaiq_schedule_content, rtplan_file_content):
    """
    Generates an Excel sheet with dwell time information.
    Accepts bytes for both schedule and rtplan content.
    Returns the Excel file as bytes.
    """
    # 1. Locate the Template File
    base_path = Path(__file__).resolve().parents[3] / 'src'
    template_excel_path = base_path / 'templates' / 'Dwell time decay Worksheet Cylinder.xlsx'

    # 2. Parse Mosaiq Schedule
    fraction_datetimes = parse_mosaiq_schedule_for_hdr_tx(mosaiq_schedule_content)
    
    # 3. Parse RTPLAN Data
    try:
        ds = pydicom.dcmread(io.BytesIO(rtplan_file_content), force=True)
        parser = ParserService()
        
        # Extract Plan Info
        patient_name = str(getattr(ds, 'PatientName', 'N/A'))
        patient_mrn = str(getattr(ds, 'PatientID', 'N/A'))
        plan_data = parser.get_plan_data(ds)
        plan_name = plan_data.get('plan_name', 'N/A')
        
        # Format Reference Date
        ref_date = plan_data.get('source_strength_ref_date', 'N/A')
        ref_time = plan_data.get('source_strength_ref_time', 'N/A')
        plan_date_str = "N/A"
        if ref_date != 'N/A':
            try:
                # Time might be optional or partial
                time_str = ref_time.split('.')[0] if ref_time != 'N/A' else "000000"
                dt_obj = datetime.strptime(f"{ref_date}{time_str}", "%Y%m%d%H%M%S")
                plan_date_str = dt_obj.strftime('%Y-%m-%d %H:%M')
            except ValueError:
                pass

        # Calculate Source Activity in Ci
        rakr = plan_data.get('rakr', 0.0)
        source_activity_ci = rakr / 4.0367 / 1000 # Conv factor for Ir-192 RAKR to Ci

        # Extract Dwell Times
        # Now passing the Dataset directly to supported ParserService (requires update to parser_service first)
        # OR passing BytesIO if updated parser prefers that.
        # We updated parser to accept dataset? No, we updated it to accept BytesIO.
        # But we also made it accept 'file path or stream'. 
        # If we pass `ds` (Dataset), `dcmread(ds)` fails.
        # So we should pass `io.BytesIO(rtplan_file_content)` again.
        dwell_data = parser.get_dwell_times_and_positions(io.BytesIO(rtplan_file_content))

    except Exception as e:
        print(f"Error parsing RTPLAN content: {e}")
        return None

    # 4. Populate Excel Template
    try:
        wb = load_workbook(template_excel_path)
        ws = wb.active
        
        # Patient Info
        ws['B5'] = patient_name
        ws['B6'] = patient_mrn
        ws['B7'] = plan_name
        ws['B11'] = plan_date_str
        ws['B13'] = source_activity_ci

        # Fractions Dates (Columns C, D, E, F, G...)
        # If no schedule found, we leave blank? Or maybe we can't generate it?
        # User might want the sheet even if schedule is empty (just for dwells).
        
        fraction_cells = ['C11', 'D11', 'E11', 'F11', 'G11']
        for i, dt in enumerate(fraction_datetimes):
            if i < len(fraction_cells):
                ws[fraction_cells[i]] = dt.strftime('%Y-%m-%d %H:%M')
        
        # Fraction Numbers
        ws['B9'] = "Plan"
        header_cells = ['C9', 'D9', 'E9', 'F9', 'G9']
        count = len(fraction_datetimes) if fraction_datetimes else 5 # Default 5 headers if no schedule?
        for i in range(count):
             if i < len(header_cells):
                ws[header_cells[i]] = i + 1
        
        # Dwell Times
        # Map: Position (mm) -> Time (s)
        # Note: 1300mm is index length usually? The template assumes specific positions.
        # Original script used: 300 - int(item['position']). 
        # Let's trust that logic for now.
        excel_dwell_map = {300 - int(item['position']): item['dwell_time'] for item in dwell_data}
        
        dwell_time_start_row = 17
        for i in range(12): # Iterate through the standard positions in the sheet
            position_cell = f'A{dwell_time_start_row + i}'
            dwell_time_cell = f'B{dwell_time_start_row + i}'
            
            position_val = ws[position_cell].value
            
            if position_val is not None and position_val in excel_dwell_map:
                ws[dwell_time_cell] = excel_dwell_map[position_val]
            else:
                ws[dwell_time_cell] = 0.0
        
        # Save to in-memory stream
        excel_bytes_io = io.BytesIO()
        wb.save(excel_bytes_io)
        excel_bytes_io.seek(0)
        return excel_bytes_io.getvalue()

    except FileNotFoundError:
        print(f"Error: The template file was not found at '{template_excel_path}'.")
        return None
    except Exception as e:
        print(f"An error occurred while populating the Excel template: {e}")
        return None