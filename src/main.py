import sys
import base64
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import pydicom
from .html_parser import parse_html_report
from .dicom_parser import find_dicom_file, load_dicom_file, get_structure_data, get_plan_data, get_dwell_times_and_positions, get_dose_data
from .calculations import get_dvh, evaluate_constraints, calculate_dose_to_meet_constraint, calculate_point_dose_bed_eqd2, get_dose_at_point, check_plan_time, calculate_bed_and_eqd2
import argparse
from pathlib import Path
import json
import os
import pdfkit

def replace_css_variables(html_content):
    """Replaces CSS variables with their actual values for PDF generation."""
    # This function remains unchanged
    colors = {
        '--text-color': '#333',
        '--background-color': '#fff',
        '--header-color-1': '#2a7ae2',
        '--header-color-2': '#1e5aab',
        '--border-color': '#ddd',
        '--table-header-bg': '#1e5aab',
        '--table-header-text': 'white',
        '--table-even-row-bg': '#eaf2fa',
        '--met-bg': '#77dd77',
        '--met-text': 'white',
        '--not-met-bg': '#ff6961',
        '--not-met-text': 'white',
        '--warning-bg': '#fdfd96',
        '--warning-text': 'black',
    }
    for var, value in colors.items():
        html_content = html_content.replace(f'var({var})', value)
    return html_content

def convert_html_to_pdf(html_content, output_path):
    """
    Converts HTML content to a PDF file using pdfkit.
    """
    # This function remains unchanged
    try:
        options = {'enable-local-file-access': None}
        if getattr(sys, 'frozen', False):
            base_path = Path(sys._MEIPASS)
        else:
            base_path = Path(os.path.dirname(os.path.abspath(__file__)))
        path_wkhtmltopdf = base_path / 'vendor' / 'bin' / 'wkhtmltopdf.exe'
        if not path_wkhtmltopdf.is_file():
            raise IOError(f"wkhtmltopdf.exe not found at the expected path: {path_wkhtmltopdf}")
        config = pdfkit.configuration(wkhtmltopdf=str(path_wkhtmltopdf))
        pdf_html_content = replace_css_variables(html_content)
        pdfkit.from_string(pdf_html_content, output_path, options=options, configuration=config)
    except IOError as e:
        if 'wkhtmltopdf' in str(e):
             raise IOError(
                "Could not generate PDF. There might be an issue with the wkhtmltopdf executable."
                f"\nPath being used: {path_wkhtmltopdf}"
                f"\n\nOriginal error: {e}"
             )
        raise e

def generate_html_report(patient_name, patient_mrn, plan_name, plan_date, plan_time, source_info, brachy_dose_per_fraction, number_of_fractions, ebrt_dose, ebrt_fractions, dvh_results, constraint_evaluation, dose_references, point_dose_results, output_path, alpha_beta_ratios, previous_brachy_data=None):
    # This function remains unchanged
    if getattr(sys, 'frozen', False):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).parent
    template_path = base_path / "templates" / "report_template.html"
    logo_path = base_path / "assets" / "2020-flame-red-02.PNG"
    with open(template_path, "r") as f:
        template = f.read()
    try:
        with open(logo_path, "rb") as img_file:
            logo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
            logo_data_uri = f"data:image/png;base64,{logo_base64}"
    except FileNotFoundError:
        logo_data_uri = ""
    previous_fractions = 0
    if previous_brachy_data and isinstance(previous_brachy_data, dict):
        max_len = 0
        if previous_brachy_data.get("dvh_results"):
            for organ_data in previous_brachy_data["dvh_results"].values():
                for dose_list in organ_data.get("dose_fx", {}).values():
                    if isinstance(dose_list, list):
                        max_len = max(max_len, len(dose_list))
        if previous_brachy_data.get("point_dose_results"):
            for dose_list in previous_brachy_data["point_dose_results"].values():
                 if isinstance(dose_list, list):
                        max_len = max(max_len, len(dose_list))
        previous_fractions = max_len
    total_fractions = previous_fractions + number_of_fractions
    fraction_headers = "".join([f"<th>Fx {i+1} Dose (Gy)</th>" for i in range(total_fractions)])
    target_volume_rows = ""
    oar_rows = ""
    for organ, data in dvh_results.items():
        alpha_beta = alpha_beta_ratios.get(organ, alpha_beta_ratios["Default"])
        volume_cc = data.get("volume_cc", "N/A")
        if isinstance(volume_cc, (int, float)):
            volume_cc = f"{volume_cc:.2f}"
        is_target = alpha_beta == 10
        if is_target:
            metrics = [
                {'name': 'D98', 'dose_key': 'd98_gy_per_fraction', 'eqd2_key': 'eqd2_d98'},
                {'name': 'D90', 'dose_key': 'd90_gy_per_fraction', 'eqd2_key': 'eqd2_d90'}
            ]
            html_rows_for_organ = []
            for i, metric in enumerate(metrics):
                fx_doses_html = ""
                if previous_brachy_data and isinstance(previous_brachy_data, dict):
                    prev_doses = previous_brachy_data.get("dvh_results", {}).get(organ, {}).get("dose_fx", {})
                    dose_list = prev_doses.get(metric['dose_key'], [])
                    if isinstance(dose_list, list):
                        fx_doses_html += "".join([f"<td>{dose:.2f}</td>" for dose in dose_list])
                current_dose = data.get(metric['dose_key'], 0)
                fx_doses_html += f'<td>{current_dose:.2f}</td>' * number_of_fractions
                eqd2_val = data.get(metric['eqd2_key'], 0)
                if i == 0:
                    html_rows_for_organ.append(
                        f'<tr>'
                        f'<td rowspan="{len(metrics)}">{organ}</td>'
                        f'<td rowspan="{len(metrics)}">{alpha_beta}</td>'
                        f'<td rowspan="{len(metrics)}">{volume_cc}</td>'
                        f'<td>{metric["name"]}</td>'
                        f'{fx_doses_html}'
                        f'<td>{eqd2_val:.2f}</td>'
                        f'</tr>'
                    )
                else:
                    html_rows_for_organ.append(
                        f'<tr>'
                        f'<td>{metric["name"]}</td>'
                        f'{fx_doses_html}'
                        f'<td>{eqd2_val:.2f}</td>'
                        f'</tr>'
                    )
            target_volume_rows += "".join(html_rows_for_organ)
        else:
            metrics = [
                {'name': 'D0.1cc', 'dose_key': 'd0_1cc_gy_per_fraction', 'eqd2_key': 'eqd2_d0_1cc'},
                {'name': 'D1cc', 'dose_key': 'd1cc_gy_per_fraction', 'eqd2_key': 'eqd2_d1cc'},
                {'name': 'D2cc', 'dose_key': 'd2cc_gy_per_fraction', 'eqd2_key': 'eqd2_d2cc'}
            ]
            html_rows_for_organ = []
            for i, metric in enumerate(metrics):
                fx_doses_html = ""
                if previous_brachy_data and isinstance(previous_brachy_data, dict):
                    prev_doses = previous_brachy_data.get("dvh_results", {}).get(organ, {}).get("dose_fx", {})
                    dose_list = prev_doses.get(metric['dose_key'], [])
                    if isinstance(dose_list, list):
                        fx_doses_html += "".join([f"<td>{dose:.2f}</td>" for dose in dose_list])
                current_dose = data.get(metric['dose_key'], 0)
                fx_doses_html += f'<td>{current_dose:.2f}</td>' * number_of_fractions
                eqd2_val = data.get(metric['eqd2_key'], 0)
                if i == 0:
                    html_rows_for_organ.append(
                        f'<tr>'
                        f'<td rowspan="{len(metrics)}">{organ}</td>'
                        f'<td rowspan="{len(metrics)}">{alpha_beta}</td>'
                        f'<td rowspan="{len(metrics)}">{volume_cc}</td>'
                        f'<td>{metric["name"]}</td>'
                        f'{fx_doses_html}'
                        f'<td>{eqd2_val:.2f}</td>'
                        f'</tr>'
                    )
                else:
                    html_rows_for_organ.append(
                        f'<tr>'
                        f'<td>{metric["name"]}</td>'
                        f'{fx_doses_html}'
                        f'<td>{eqd2_val:.2f}</td>'
                        f'</tr>'
                    )
            oar_rows += "".join(html_rows_for_organ)
    point_dose_rows = ""
    for pr in point_dose_results:
        point_fraction_cells = ""
        if previous_brachy_data and isinstance(previous_brachy_data, dict):
            prev_doses_list = previous_brachy_data.get("point_dose_results", {}).get(pr.get('name', ''), [])
            if isinstance(prev_doses_list, list):
                for dose in prev_doses_list:
                    point_fraction_cells += f"<td>{dose:.2f}</td>"
        point_fraction_cells += f'<td>{pr.get("dose", 0):.2f}</td>' * number_of_fractions
        point_alpha_beta = alpha_beta_ratios.get(pr.get('name', 'Default'), alpha_beta_ratios["Default"])
        point_dose_rows += (
            f'<tr>'
            f'<td>{pr.get("name", "N/A")}</td>'
            f'<td>{point_alpha_beta}</td>'
            f'{point_fraction_cells}'
            f'<td>{pr.get("EQD2", 0):.2f}</td>'
            f'</tr>'
        )
    html_content = template.replace("{{ patient_name }}", patient_name)
    html_content = html_content.replace("{{ patient_mrn }}", patient_mrn)
    html_content = html_content.replace("{{ plan_name }}", plan_name)
    html_content = html_content.replace("{{ plan_date }}", plan_date)
    html_content = html_content.replace("{{ plan_time }}", plan_time)
    html_content = html_content.replace("{{ source_info }}", source_info)
    html_content = html_content.replace("{{ brachy_dose_per_fraction }}", str(brachy_dose_per_fraction))
    html_content = html_content.replace("{{ number_of_fractions }}", str(number_of_fractions))
    html_content = html_content.replace("{{ ebrt_dose }}", str(ebrt_dose))
    html_content = html_content.replace("{{ ebrt_fractions }}", str(ebrt_fractions))
    html_content = html_content.replace("{{ target_volume_rows }}", target_volume_rows)
    html_content = html_content.replace("{{ oar_rows }}", oar_rows)
    html_content = html_content.replace("{{ logo_base64 }}", logo_data_uri)
    html_content = html_content.replace("{{ fraction_headers }}", fraction_headers)
    html_content = html_content.replace("{{ point_dose_rows }}", point_dose_rows)
    with open(output_path, "w", encoding='utf-8') as f:
        f.write(html_content)
    return html_content

from fuzzywuzzy import process

def get_structure_mapping(current_structures, json_structures):
    # This function remains unchanged
    mapping = {}
    for current_struct in current_structures:
        match, score = process.extractOne(current_struct, json_structures)
        if score > 80:
            mapping[current_struct] = match
    return mapping

def main(args, structure_data, plan_data, selected_point_names=None, custom_constraints=None, dose_point_mapping=None, num_fractions_delivered=None, ebrt_fractions=None, structure_mapping=None, confirmed_structure_mapping=None):
    data_dir = Path(args.data_dir)

    if hasattr(args, 'alpha_beta_ratios') and args.alpha_beta_ratios:
        current_alpha_beta_ratios = args.alpha_beta_ratios.copy()
    else:
        from .config import templates
        current_alpha_beta_ratios = templates["Cervix HDR - EMBRACE II"]["alpha_beta_ratios"].copy()
    
    if "Default" not in current_alpha_beta_ratios:
        from .config import templates
        current_alpha_beta_ratios["Default"] = templates["Cervix HDR - EMBRACE II"]["alpha_beta_ratios"]["Default"]

    point_dose_results, previous_brachy_bed_per_organ = [], {}

    if hasattr(args, 'previous_brachy_data') and args.previous_brachy_data:
            if isinstance(args.previous_brachy_data, dict):
                # Correctly parse DVH results by accessing the nested 'dose_fx' dictionary
                for organ, data in args.previous_brachy_data.get('dvh_results', {}).items():
                    organ_lower = organ.lower()
                    if organ not in previous_brachy_bed_per_organ:
                        previous_brachy_bed_per_organ[organ] = {}
                    alpha_beta = current_alpha_beta_ratios.get(organ, current_alpha_beta_ratios["Default"])
                    
                    dose_fx_dict = data.get('dose_fx', {}) # Access the inner dictionary
                    for metric, dose_list in dose_fx_dict.items():
                        simple_metric = metric.replace('_gy_per_fraction', '')
                        if isinstance(dose_list, list):
                            total_metric_bed = sum(dose * (1 + dose / alpha_beta) for dose in dose_list)
                            previous_brachy_bed_per_organ[organ][simple_metric] = total_metric_bed

                # Correctly parse point dose results
                for point_name, data in args.previous_brachy_data.get('point_dose_results', {}).items():
                    alpha_beta = current_alpha_beta_ratios.get(point_name, current_alpha_beta_ratios["Default"])
                    dose_list = data.get('dose_fx', []) # Get the dose list for the point
                    if isinstance(dose_list, list):
                        total_point_bed = sum(dose * (1 + dose / alpha_beta) for dose in dose_list)
                        previous_brachy_bed_per_organ[point_name] = total_point_bed
        # --- START: corrected EQD2 with json brachy data ---

    dose_dir = next((d for d in Path(args.data_dir).iterdir() if d.is_dir() and "RTDOSE" in d.name), None)
    struct_dir = next((d for d in Path(args.data_dir).iterdir() if d.is_dir() and "RTst" in d.name), None)
    dose_file, struct_file = find_dicom_file(dose_dir), find_dicom_file(struct_dir)

    planned_number_of_fractions = plan_data.get('number_of_fractions', 1)
    number_of_fractions_for_calc = num_fractions_delivered if num_fractions_delivered is not None else planned_number_of_fractions

    alpha_beta_for_calc = {}
    if structure_mapping:
        for dicom_name, mapped_name in structure_mapping.items():
            standard_ab_ratio = current_alpha_beta_ratios.get(mapped_name.capitalize(), current_alpha_beta_ratios.get(mapped_name, current_alpha_beta_ratios["Default"]))
            alpha_beta_for_calc[dicom_name] = standard_ab_ratio
    else:
        alpha_beta_for_calc = current_alpha_beta_ratios

    dvh_results_orig_names = get_dvh(
        struct_file, dose_file, structure_data, number_of_fractions_for_calc,
        ebrt_dose=args.ebrt_dose, ebrt_fractions=ebrt_fractions,
        previous_brachy_bed_per_organ=previous_brachy_bed_per_organ,
        alpha_beta_ratios=alpha_beta_for_calc,
        confirmed_structure_mapping=confirmed_structure_mapping  # <-- ADD THIS LINE
    )
    
    # --- START: FINAL Case-Insensitive Mapping Logic ---
    dvh_results_mapped = {}
    if structure_mapping:
        case_insensitive_mapping = {key.lower(): value for key, value in structure_mapping.items()}
        for original_dicom_name, dvh_data in dvh_results_orig_names.items():
            from .calculations import normalize_structure_name
            normalized_lookup_key = normalize_structure_name(original_dicom_name)
            
            mapped_name = case_insensitive_mapping.get(normalized_lookup_key.lower())
            
            if mapped_name and mapped_name != "Ignore":
                dvh_results_mapped[mapped_name] = dvh_data
            # Also pass through unmapped OARs
            elif original_dicom_name in ["Bladder", "Rectum", "Sigmoid", "Bowel"]:
                 dvh_results_mapped[original_dicom_name] = dvh_data
    else:
        # Fallback if no mapping was provided
        dvh_results_mapped = dvh_results_orig_names
    # --- END: FINAL Case-Insensitive Mapping Logic ---

    for organ, data in dvh_results_mapped.items():

        dose_metrics = {
            'd2cc': 'd2cc_gy_per_fraction', 'd1cc': 'd1cc_gy_per_fraction', 'd0_1cc': 'd0_1cc_gy_per_fraction',
            'd90': 'd90_gy_per_fraction', 'd98': 'd98_gy_per_fraction', 'd95': 'd95_gy_per_fraction',
            'max': 'max_dose_gy_per_fraction', 'mean': 'mean_dose_gy_per_fraction', 'min': 'min_dose_gy_per_fraction',
        }
        for metric_key, dose_key in dose_metrics.items():
            dose_per_fraction = data.get(dose_key, 0)
            # This now correctly calculates the dose for the NEW fractions planned
            total_dose = dose_per_fraction * number_of_fractions_for_calc
            previous_brachy_bed = data.get('previous_brachy_bed', {}).get(metric_key, 0)
            total_bed, eqd2, bed_brachy, _, _ = calculate_bed_and_eqd2(
                total_dose, dose_per_fraction, organ, args.ebrt_dose, ebrt_fractions,
                previous_brachy_bed, current_alpha_beta_ratios
            )
            data[f'bed_{metric_key}'], data[f'eqd2_{metric_key}'], data[f'bed_brachy_{metric_key}'] = total_bed, eqd2, bed_brachy

    current_target_constraints = custom_constraints.get("constraints", {}).get("target_constraints", {})
    current_oar_constraints = custom_constraints.get("constraints", {}).get("oar_constraints", {})
    point_dose_constraints = custom_constraints.get("point_dose_constraints", {})
    
    filtered_dose_references = [dr for dr in plan_data.get('dose_references', []) if (selected_point_names and dr['name'] in selected_point_names) or not selected_point_names]
    for dr in filtered_dose_references:
        total_bed, eqd2, bed_brachy, bed_ebrt, bed_previous_brachy = calculate_point_dose_bed_eqd2(
            dr['dose'], number_of_fractions_for_calc, dr['name'], args.ebrt_dose, ebrt_fractions,
            previous_brachy_bed=previous_brachy_bed_per_organ.get(dr['name'], 0), alpha_beta_ratios=current_alpha_beta_ratios
        )
        point_dose_results.append({
            'name': dr['name'], 'dose': dr['dose'], 'total_dose': dr['dose'] * number_of_fractions_for_calc,
            'BED_this_plan': bed_brachy, 'BED_previous_brachy': bed_previous_brachy, 'BED_EBRT': bed_ebrt, 'EQD2': eqd2,
        })

    constraint_evaluation = evaluate_constraints(dvh_results_mapped, point_dose_results, current_target_constraints, current_oar_constraints, point_dose_constraints, dose_point_mapping)

    mapping_dict = {item[0]: item[1] for item in dose_point_mapping} if dose_point_mapping else {}
    for pr in point_dose_results:
        status_updated = False
        mapped_constraint_name = mapping_dict.get(pr['name'])
        if mapped_constraint_name and mapped_constraint_name in point_dose_constraints:
            constraint = point_dose_constraints[mapped_constraint_name]
            if constraint.get("check_type") == "prescription_tolerance":
                tolerance = constraint.get("tolerance", 0.0)
                prescribed_dose = plan_data.get('brachy_dose_per_fraction', 0)
                if prescribed_dose > 0:
                    pr['Constraint Status'] = 'Pass' if prescribed_dose * (1 - tolerance) <= pr['dose'] <= prescribed_dose * (1 + tolerance) else 'Fail'
                    status_updated = True
            elif "max_eqd2" in constraint:
                pr['Constraint Status'] = 'Pass' if pr['EQD2'] <= constraint["max_eqd2"] else 'Fail'
                status_updated = True
        if not status_updated:
            point_eval = constraint_evaluation.get(f"Point Dose - {pr['name']}", {})
            pr['Constraint Status'] = point_eval.get('status', 'N/A')

    for organ, data in dvh_results_mapped.items():
        if organ in constraint_evaluation and constraint_evaluation[organ].get("EQD2_met") == "False":
            eqd2_constraint = constraint_evaluation[organ]["EQD2_max"]
            prev_bed_dict = previous_brachy_bed_per_organ.get(organ, {})
            prev_d2cc_bed = prev_bed_dict.get('d2cc', 0) if isinstance(prev_bed_dict, dict) else 0
            data["dose_to_meet_constraint"] = calculate_dose_to_meet_constraint(
                eqd2_constraint, organ, number_of_fractions_for_calc, args.ebrt_dose,
                ebrt_fractions=ebrt_fractions, previous_brachy_bed=prev_d2cc_bed, alpha_beta_ratios=current_alpha_beta_ratios
            )
        else:
            data["dose_to_meet_constraint"] = "N/A"

    source_strength_ref_date = plan_data.get('source_strength_ref_date', 'N/A')
    source_strength_ref_time = plan_data.get('source_strength_ref_time', 'N/A')
    if source_strength_ref_date != 'N/A' and source_strength_ref_time != 'N/A':
        plan_datetime = datetime.strptime(f"{source_strength_ref_date}{source_strength_ref_time.split('.')[0]}", "%Y%m%d%H%M%S")
        formatted_plan_date, formatted_plan_time = plan_datetime.strftime('%Y-%m-%d'), plan_datetime.strftime('%H:%M:%S')
    else:
        formatted_plan_date, formatted_plan_time = 'N/A', 'N/A'
    
    rt_dose_dataset = load_dicom_file(dose_file)
    
    output_data = {
        "patient_name": str(rt_dose_dataset.PatientName), "patient_mrn": str(rt_dose_dataset.PatientID),
        "plan_name": plan_data.get('plan_name', 'N/A'), "plan_date": formatted_plan_date, "plan_time": formatted_plan_time,
        "source_info": plan_data.get('source_info', 'N/A'), "channel_mapping": plan_data.get('channel_mapping', []),
        "brachy_dose_per_fraction": plan_data.get('brachy_dose_per_fraction', 0), "planned_number_of_fractions": planned_number_of_fractions,
        "calculation_number_of_fractions": number_of_fractions_for_calc, "ebrt_dose": args.ebrt_dose,
        "dvh_results": dvh_results_mapped,
        "constraint_evaluation": constraint_evaluation,
        "point_dose_results": point_dose_results,
        "plan_time_warning": check_plan_time(plan_data.get('plan_time')),
    }

    if args.output_html:
        html_content = generate_html_report(
            output_data["patient_name"], output_data["patient_mrn"], output_data["plan_name"], 
            output_data["plan_date"], output_data["plan_time"], output_data["source_info"],
            output_data["brachy_dose_per_fraction"], output_data["calculation_number_of_fractions"], 
            output_data["ebrt_dose"], ebrt_fractions, output_data["dvh_results"], 
            output_data["constraint_evaluation"], plan_data.get('dose_references', []), 
            output_data["point_dose_results"], args.output_html, current_alpha_beta_ratios,
            previous_brachy_data=args.previous_brachy_data
        )
        output_data['html_report'] = html_content

    return output_data

def generate_dwell_time_sheet(mosaiq_schedule_path, rtplan_file, output_excel_path):
    # This function remains unchanged
    def parse_mosaiq_schedule_for_hdr_tx(file_path):
        try:
            df = pd.read_excel(file_path)
            hdr_tx_schedule = df[
                df['Activity'].str.contains('HDR', case=False, na=False) &
                df['Description'].str.contains('tx', case=False, na=False)
            ].copy()
            hdr_tx_schedule = hdr_tx_schedule[~hdr_tx_schedule['Sts'].str.contains('X', na=False)]
            hdr_tx_schedule['Date'] = pd.to_datetime(hdr_tx_schedule['Date'], errors='coerce')
            hdr_tx_schedule['Time'] = hdr_tx_schedule['Time'].astype(str)
            hdr_tx_schedule.dropna(subset=['Date'], inplace=True)
            hdr_tx_schedule['datetime'] = pd.to_datetime(
                hdr_tx_schedule['Date'].dt.strftime('%Y-%m-%d') + ' ' + hdr_tx_schedule['Time'],
                format='mixed'
            )
            return sorted(hdr_tx_schedule['datetime'].tolist())
        except Exception as e:
            print(f"Error parsing Mosaiq schedule file: {e}")
            return []
    if getattr(sys, 'frozen', False):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).parent.parent
    template_excel_path = base_path / 'src' / 'templates' / 'Dwell time decay Worksheet Cylinder.xlsx'
    fraction_datetimes = parse_mosaiq_schedule_for_hdr_tx(mosaiq_schedule_path)
    if not fraction_datetimes:
        print("No 'HDR: tx' activities found in the schedule. Exiting.")
        return
    patient_name = "N/A"
    patient_mrn = "N/A"
    plan_name = "N/A"
    plan_date_str = "N/A"
    if rtplan_file:
        rtplan_dataset = pydicom.dcmread(rtplan_file)
        patient_name = str(rtplan_dataset.PatientName)
        patient_mrn = str(rtplan_dataset.PatientID)
        plan_data = get_plan_data(rtplan_file)
        plan_name = plan_data.get('plan_name', 'N/A')
        source_strength_ref_date = plan_data.get('source_strength_ref_date', 'N/A')
        source_strength_ref_time = plan_data.get('source_strength_ref_time', 'N/A')
        if source_strength_ref_date != 'N/A' and source_strength_ref_time != 'N/A':
            plan_datetime = datetime.strptime(f"{source_strength_ref_date}{source_strength_ref_time.split('.')[0]}", "%Y%m%d%H%M%S")
            plan_date_str = plan_datetime.strftime('%Y-%m-%d %H:%M')
        else:
            plan_date_str = "N/A"
        rakr = plan_data.get('rakr', 0.0)
        source_activity_ci = rakr / 4.0367 / 1000
    else:
        source_activity_ci = 0.0
    try:
        wb = load_workbook(template_excel_path)
        ws = wb.active
        ws['B5'] = patient_name
        ws['B6'] = patient_mrn
        ws['B7'] = plan_name
        ws['B11'] = plan_date_str
        fraction_cells = ['C11', 'D11', 'E11', 'F11', 'G11']
        for i, dt in enumerate(fraction_datetimes):
            if i < len(fraction_cells):
                ws[fraction_cells[i]] = dt.strftime('%Y-%m-%d %H:%M')
        ws['B9'] = "Plan"
        header_cells = ['C9', 'D9', 'E9', 'F9', 'G9']
        for i in range(len(fraction_datetimes)):
             if i < len(header_cells):
                ws[header_cells[i]] = i + 1
        ws['B13'] = source_activity_ci
        dwell_data = get_dwell_times_and_positions(rtplan_file)
        excel_dwell_map = {300 - int(item['position']): item['dwell_time'] for item in dwell_data}
        dwell_time_start_row = 17
        for i in range(12):
            position_cell = f'A{dwell_time_start_row + i}'
            dwell_time_cell = f'B{dwell_time_start_row + i}'
            position = ws[position_cell].value
            if position is not None and position in excel_dwell_map:
                ws[dwell_time_cell] = excel_dwell_map[position]
            else:
                ws[dwell_time_cell] = 0.0
        wb.save(output_excel_path)
    except FileNotFoundError:
        print(f"Error: The template file was not found at '{template_excel_path}'.")
    except Exception as e:
        print(f"An error occurred while populating the Excel template: {e}")